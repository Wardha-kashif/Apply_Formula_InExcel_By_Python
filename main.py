import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

wb = load_workbook('sheet1.xlsx')
sheet = wb['Sheet1']
# cell references (original spreadsheet)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

print(min_column)
print(max_column)
print(min_row)
print(max_row)


import string
alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column]
print(excel_alphabet)


for i in excel_alphabet:
    if i=='G':
        for j in range(max_row):
            sheet[f'{i}{j + 2}']= f'=(C{j + 2}/100 * 100)'

wb.save('sheet1.xlsx')


# adding total label
# sheet[f'{excel_alphabet[0]}{max_column+1}'] = 'Total'
# wb.save('sheet1.xlsx')

